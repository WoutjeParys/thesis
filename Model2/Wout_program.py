__author__ = 'Wout'

#import some usefull things
import os
from pandas import pivot_table, merge, ExcelWriter, DataFrame
import numpy as np
import openpyxl
from gams_addon import gdx_to_df, DomainInfo
from openpyxl.styles import Style, Border, Alignment, Protection, Font, colors
import sqlite3 as sq
import csv
import xlrd
import os
import math
from matplotlib import pyplot as plt
import Wout_initialise
import Wout_main



file = 'results\out_db_75.6_50.gdx'
gdx_file = os.path.join(os.getcwd(), '%s' % file)
sh_shift_name = 'shifting'
sh_ratio_name = 'ratio'
excel_shift_name   = 'excel\output_elasticity_model_shifting.xlsx'
excel_tempory_name = 'excel\output_elasticity_model_tempory.xlsx'
list_compensation = list()
list_ratio = list()

#length period needs to be a multiple of 24
#percentageEV lies in the range 0-100
length_period = 24*7
amount_of_periods = 8
startday_weekend = 2
percentageEV = 100
season_range = 1

#reset compensation factor for inbalances in elasticities (to 1)
def reset_ratio():
    resetvalue = 0.05
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()

    print os.getcwd()
    sql = 'DROP TABLE IF EXISTS Ratio;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Ratio (Period TEXT, Hour TEXT, Value FLOAT);'
    # ,  PRIMARY KEY(Code));'
    cur.execute(sql)
    ratios = list()
    for p in range(1,amount_of_periods+1):
        ratios_period = list()
        for h in range(1,length_period+1):
            ratios.append((p,h,resetvalue))
            ratios_period.append(resetvalue)
            # print (p,' & ', h)
        print ratios_period
        list_ratio.append(ratios_period)
    cur.executemany('INSERT INTO Ratio VALUES (?,?,?)', ratios)
    conn.commit()

    ############################################

# set ratio based on a given file
def set_ratio(gdx_file):
    # gdx_file = 'results\8weeksFull\out_db_5_DR'
    #define the used files
    writefile = os.getcwd() + '\\' + 'excel\output_elasticity_model_tempory.xlsx'
    writer = ExcelWriter(writefile)
    file = gdx_file
    gdx_file = os.path.join(os.getcwd(), '%s' % file)
    print gdx_file
    zone_dict = dict()
    zone_dict['BEL_Z'] = 'BEL'

    #gdx to excel
    print 'Retrieving ratio'
    ratio = gdx_to_df(gdx_file, 'ratio')
    old_index = ratio.index.names
    ratio['C'] = [zone_dict[z] for z in ratio.index.get_level_values('Z')]
    ratio.set_index('C', append=True, inplace=True)
    ratio = ratio.reorder_levels(['C'] + old_index)
    ratio.reset_index(inplace=True)
    ratio = pivot_table(ratio, 'ratio', index=['P','H','Z'], columns=['C'], aggfunc=np.sum)
    print 'Writing ratio to Excel'
    ratio.to_excel(writer, na_rep=0.0, sheet_name='ratio', merge_cells=False)
    writer.close()

    #calculate new ratios and put in sql
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()
    print os.getcwd()

    sql = 'DROP TABLE IF EXISTS Ratio;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Ratio (Period TEXT, Hour TEXT, Value FLOAT);'
    cur.execute(sql)
    ratios = list()
    print 'open tempory excel file'
    wbread = openpyxl.load_workbook(excel_tempory_name)
    print 'tempory file loaded'
    sheet = wbread.get_sheet_by_name(sh_ratio_name)
    for i in range(1,amount_of_periods+1):
        for j in range(2,length_period+2):
            list_ratio[i-1][j-2] = round(list_ratio[i-1][j-2]*(sheet.cell(row = (i-1)*length_period+j,column = 4).value),4)
            # print 'list_ratio for i = ', i, ', and j = ', j
            # print list_ratio[i-1][j-2]
            period = i
            hour = j-1
            ratio = list_ratio[i-1][j-2]
            # print period, hour, ratio
            ratios.append((period,hour,ratio))
    cur.executemany('INSERT INTO Ratio VALUES (?,?,?)', ratios)
    conn.commit()

    print 'done ratio'

#get the inbalance ratio for each hour h
def set_inbalance_ratio():
    #define the used files
    writefile = os.getcwd() + '\\' + 'excel\output_elasticity_model_tempory.xlsx'
    writer = ExcelWriter(writefile)
    print gdx_file
    zone_dict = dict()
    zone_dict['BEL_Z'] = 'BEL'

    #gdx to excel
    print 'Retrieving ratio'
    ratio = gdx_to_df(gdx_file, 'ratio')
    old_index = ratio.index.names
    ratio['C'] = [zone_dict[z] for z in ratio.index.get_level_values('Z')]
    ratio.set_index('C', append=True, inplace=True)
    ratio = ratio.reorder_levels(['C'] + old_index)
    ratio.reset_index(inplace=True)
    ratio = pivot_table(ratio, 'ratio', index=['P','H','Z'], columns=['C'], aggfunc=np.sum)
    print 'Writing ratio to Excel'
    ratio.to_excel(writer, na_rep=0.0, sheet_name='ratio', merge_cells=False)
    writer.close()

    #calculate new ratios and put in sql
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()
    print os.getcwd()

    sql = 'DROP TABLE IF EXISTS Ratio;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Ratio (Period TEXT, Hour TEXT, Value FLOAT);'
    cur.execute(sql)
    ratios = list()
    print 'open tempory excel file'
    wbread = openpyxl.load_workbook(excel_tempory_name)
    print 'tempory file loaded'
    sheet = wbread.get_sheet_by_name(sh_ratio_name)
    for i in range(1,amount_of_periods+1):
        for j in range(2,length_period+2):
            list_ratio[i-1][j-2] = round(list_ratio[i-1][j-2]*(sheet.cell(row = (i-1)*length_period+j,column = 4).value),4)
            # print 'list_ratio for i = ', i, ', and j = ', j
            # print list_ratio[i-1][j-2]
            period = i
            hour = j-1
            ratio = list_ratio[i-1][j-2]
            # print period, hour, ratio
            ratios.append((period,hour,ratio))
    cur.executemany('INSERT INTO Ratio VALUES (?,?,?)', ratios)
    conn.commit()

    print 'done ratio'

#Auxiliary funtion to calculate power of negative functions
def calculate_power(value, power):
    if value < 0:
        result = -math.pow(-value,power)
    else:
        result = math.pow(value,power)
    return result

# function to set demand profiles for different penetration rates
def set_demandprofiles(case):

    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()
    print os.getcwd()

    file = "excel\DemR\penetrationDR\DemAllProfiles0_" + str(case) + ".xlsx"

    book = xlrd.open_workbook(os.path.join(os.getcwd() , file))
    sqlref = 'DROP TABLE IF EXISTS Dem_ref_profile;'
    sqlmin = 'DROP TABLE IF EXISTS Dem_min_profile;'
    sqlmax = 'DROP TABLE IF EXISTS Dem_max_profile;'
    sqlflat = 'DROP TABLE IF EXISTS Dem_flat_profile;'
    cur.execute(sqlref)
    cur.execute(sqlmin)
    cur.execute(sqlflat)
    cur.execute(sqlmax)
    sqlref = 'CREATE TABLE IF NOT EXISTS Dem_ref_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    sqlmin = 'CREATE TABLE IF NOT EXISTS Dem_min_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    sqlmax = 'CREATE TABLE IF NOT EXISTS Dem_max_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    sqlflat = 'CREATE TABLE IF NOT EXISTS Dem_flat_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    cur.execute(sqlref)
    cur.execute(sqlmin)
    cur.execute(sqlmax)
    cur.execute(sqlflat)
    demref = list()
    demmin = list()
    demmax = list()
    demflp = list()
    zone = 'BEL_Z'
    shmin=book.sheet_by_index(0)
    shmax=book.sheet_by_index(1)
    shref=book.sheet_by_index(2)
    shflp=book.sheet_by_index(3)
    amount_of_days = length_period/24
    for season in range (0,4):
        print 'season: ', season
        for day in range(0,amount_of_days):
            if day == startday_weekend or day == startday_weekend+1:
                print 'weekendday with row = ', season*2+1
                row = season*2+2
            else:
                print 'weekday with sheet index = ', season*2
                row = season*2+1
            for col in range(1,shref.ncols):
                hour = int(shref.cell_value(0,col)) + 24*day
                valueref = shref.cell_value(row,col)
                valuemin = shmin.cell_value(row,col)
                valuemax = shmax.cell_value(row,col)
                valueflp = shflp.cell_value(row,col)
                demref.append((season+1,zone,hour,valueref))
                demmin.append((season+1,zone,hour,valuemin))
                demmax.append((season+1,zone,hour,valuemax))
                demflp.append((season+1,zone,hour,valueflp))
    cur.executemany('INSERT INTO Dem_ref_profile VALUES (?,?,?,?)', demref)
    cur.executemany('INSERT INTO Dem_min_profile VALUES (?,?,?,?)', demmin)
    cur.executemany('INSERT INTO Dem_max_profile VALUES (?,?,?,?)', demmax)
    cur.executemany('INSERT INTO Dem_flat_profile VALUES (?,?,?,?)', demflp)
    conn.commit()
    print 'Done price profiles'

# function to set elasticity for different penetration rates
def set_elasticity(case):
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()
    print os.getcwd()

    file = "excel\DemR\penetrationDR\Elasticity0_" + str(case) + ".xlsx"

    book = xlrd.open_workbook(os.path.join(os.getcwd() , file))
    sql = 'DROP TABLE IF EXISTS Elasticity;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Elasticity (Season FLOAT, Hour1 TEXT, Hour2 TEXT, Price_Elasticity FLOAT);'
    cur.execute(sql)
    elasticity = list()
    # TODO
    # check how to handle elasticity, for now, only Hour1 - Hour2 and matrix 168-168
    # amount of days should be 7 to work with right elasticities weekday-weekend
    amount_of_days = length_period/24
    for season in range (0,4):
        print 'season: ', season
        print 'season: ', season
        for day in range(0,amount_of_days):
            if day == startday_weekend or day == startday_weekend+1:
                print 'in if, and index = ', season*2+1
                sh=book.sheet_by_index(season*2+1)
            else:
                print 'in else, and index = ', season*2
                sh=book.sheet_by_index(season*2)
            for row in range(3,sh.nrows):
                hour1 = int(sh.cell_value(row, 0)) + 24*day
                for col in range(1,sh.ncols):
                    if col < 13:
                        if row > 14 + col:
                            hour2 = int(sh.cell_value(2, col)) + 24*(day+1)
                        else:
                            hour2 = int(sh.cell_value(2, col)) + 24*(day)
                        if hour2 > length_period:
                            hour2 = hour2 - length_period
                    else:
                        if col > 11 + row-2:
                            hour2 = int(sh.cell_value(2, col)) + 24*(day-1)
                        else:
                            hour2 = int(sh.cell_value(2, col)) + 24*(day)
                        if hour2 < 1:
                            hour2 = hour2 + length_period
                    value = sh.cell_value(row,col)
                    elasticity.append((season+1,hour1,hour2,value))
    cur.executemany('INSERT INTO Elasticity VALUES (?,?,?,?)', elasticity)
    conn.commit()
    print 'Done elasticities'

Wout_initialise.initialise(length_period)


#no DR
note = 'noDR'
string1 = 'PRICE_REF(P,H,Z) = P_REF;\n'
string2 = 'DEM_OPTIMAL(P,T,Z) = DEM_RES_FP(P,T,Z);\n'
string3 = 'LIMITPRICE = 0;\n'
string4 = 'FACTOR_RES_DR = 0;\n'
stringtot = string1+string2+string3+string4
for res_target_extern in []:
    Wout_main.main(length_period,res_target_extern,note,stringtot)

print '-------------------------'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print 'NO DEMAND RESPONSE CASES ARE DONE!'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print '-------------------------'

# with DR
reset_ratio()
set_ratio('results\8weeksFull\out_db_5_DR')

print '-------------------------'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print 'CALCULATE RATIO IS DONE FOR UPCOMING CASES!!!'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print '-------------------------'

# DR not as reserves
note = 'DR'
string3 = 'LIMITPRICE = 1.5;\n'
string4 = 'FACTOR_RES_DR = 0;\n'
stringtot = string3+string4

for res_target_extern in []:
    Wout_main.main(length_period,res_target_extern,note,stringtot)
# #     file = 'results\out_db_'+ str(res_target_extern) + '_DR.gdx'
# #     gdx_file = os.path.join(os.getcwd(), '%s' % file)
# #     set_inbalance_ratio()

print '-------------------------'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print 'CASES WITH DEMAND RESPONSE ARE DONE!!!!!!!!!'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print '-------------------------'

# # DR also as reserves
note = 'DRres'
string3 = 'LIMITPRICE = 1.5;\n'
string4 = 'FACTOR_RES_DR = 1;\n'
stringtot = string3+string4

for res_target_extern in []:
    Wout_main.main(length_period,res_target_extern,note,stringtot)
#     file = 'results\out_db_'+ str(res_target_extern) + '_DRres.gdx'
#     gdx_file = os.path.join(os.getcwd(), '%s' % file)
#     set_inbalance_ratio()

# DR also as reserves
note = 'DRres0_1'
string3 = 'LIMITPRICE = 1.5;\n'
string4 = 'FACTOR_RES_DR = 0.1;\n'
stringtot = string3+string4

for res_target_extern in [20]:
    Wout_main.main(length_period,res_target_extern,note,stringtot)

# currently only choice between 25,50 or 75
for penetration in [75]:
    set_demandprofiles(penetration)
    set_elasticity(penetration)
    note = 'DRpen_' + str(penetration)
    string3 = 'LIMITPRICE = 1.5;\n'
    string4 = 'FACTOR_RES_DR = 0.1;\n'
    stringtot = string3+string4
    for res_target_extern in [20,50]:
        Wout_main.main(length_period,res_target_extern,note,stringtot)


print '-------------------------'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print 'AND WE ARE COMPLETELY FINISHED!!!!!!!!!!'
print '\n'
print '\n'
print '\n'
print '\n'
print '\n'
print '-------------------------'


#write marginal values of balance function to excel
def balance_m_to_excel():
    writefile = os.getcwd() + '\\' + 'excel\output_elasticity_model_tempory.xlsx'

    writer = ExcelWriter(writefile)
    print gdx_file
    zone_dict = dict()
    zone_dict['BEL_Z'] = 'BEL'

    print 'get balance'
    print 'retrieving marg'
    marg = gdx_to_df(gdx_file, 'marg')
    old_index = marg.index.names
    marg['C'] = [zone_dict[z] for z in marg.index.get_level_values('Z')]
    marg.set_index('C', append=True, inplace=True)
    marg = marg.reorder_levels(['C'] + old_index)
    marg.reset_index(inplace=True)
    marg = pivot_table(marg, 'marg', index=['P', 'T', 'Z'], columns=['C'], aggfunc=np.sum)

    print 'Writing balances.m to Excel'
    marg.to_excel(writer, na_rep=0.0, sheet_name='balance', merge_cells=False)
    writer.close()

#write marginal values of balance function from excel to sqlite
def balance_m_to_sqlite():
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()

    print os.getcwd()

    book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\output_elasticity_model_tempory.xlsx"))
    sh = book.sheet_by_index(0)
    sql = 'DROP TABLE IF EXISTS Marketprices;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Marketprices (Period TEXT, Hour TEXT, Zone Text, Price FLOAT);'
    cur.execute(sql)
    prices = list()
    for row in range(1, sh.nrows):
        period = int(sh.cell_value(row, 0))
        hour = int(sh.cell_value(row,1))
        zone = sh.cell_value(row,2)
        #if zone in zones:
        price = sh.cell_value(row, 3)
        #print period, hour, zone, price
        prices.append((period,hour,zone,price))
    cur.executemany('INSERT INTO Marketprices VALUES (?,?,?,?)', prices)
    conn.commit()

    print 'done marketprices'

#reset compensation factor for inbalances in elasticities (to 1)
def reset_factor_to_one():
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()

    print os.getcwd()
    sql = 'DROP TABLE IF EXISTS Factor;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Factor (Period TEXT, Hour TEXT, Value FLOAT);'
    # ,  PRIMARY KEY(Code));'
    cur.execute(sql)
    factors = list()
    for p in range(1,amount_of_periods+1):
        for h in range(1,length_period+1):
            factors.append((p,h,1.0))
            # print (p,' & ', h)
    cur.executemany('INSERT INTO Factor VALUES (?,?,?)', factors)
    conn.commit()

    for i in range(1,amount_of_periods+1):
        list_compensation.append(1)

    ############################################

#reset compensation factor for inbalances in elasticities (to 1)
def reset_factor_to_value(value):
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()

    print os.getcwd()
    sql = 'DROP TABLE IF EXISTS Factor;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Factor (Period TEXT, Hour TEXT, Value FLOAT);'
    # ,  PRIMARY KEY(Code));'
    cur.execute(sql)
    factors = list()
    for p in range(1,amount_of_periods+1):
        for h in range(1,length_period+1):
            factors.append((p,h,value))
            # print (p,' & ', h)
    cur.executemany('INSERT INTO Factor VALUES (?,?,?)', factors)
    conn.commit()

    for i in range(1,amount_of_periods+1):
        list_compensation.append(value)

    ############################################

#set compensation factor for inbalances to a chosen value
def set_factor_to_value():
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()

    print os.getcwd()
    sql = 'DROP TABLE IF EXISTS Factor;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Factor (Period TEXT, Hour TEXT, Value FLOAT);'
    # ,  PRIMARY KEY(Code));'
    cur.execute(sql)
    factors = list()
    for p in range(1,amount_of_periods+1):
        for h in range(1,length_period+1):
            factors.append((p,h,list_compensation[p-1]))
            # print (p,' & ', h)
    cur.executemany('INSERT INTO Factor VALUES (?,?,?)', factors)
    conn.commit()
    ############################################

#update parameter to adjust cross elasticities
def update_factor_values():
    writer = ExcelWriter(writefile)
    print gdx_file
    zone_dict = dict()
    zone_dict['BEL_Z'] = 'BEL'

    print 'get compensation'
    print 'retrieving factor'
    factor = gdx_to_df(gdx_file, 'factor')
    old_index = factor.index.names
    factor['C'] = [zone_dict[z] for z in factor.index.get_level_values('Z')]
    factor.set_index('C', append=True, inplace=True)
    factor = factor.reorder_levels(['C'] + old_index)
    factor.reset_index(inplace=True)
    factor = pivot_table(factor, 'factor', index=['P', 'H', 'Z'], columns=['C'], aggfunc=np.sum)
    print 'Writing factor to Excel'
    factor.to_excel(writer, na_rep=0.0, sheet_name='factor', merge_cells=False)
    writer.close()

    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()
    print os.getcwd()

    book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\output_elasticity_model_tempory.xlsx"))
    sh = book.sheet_by_index(0)
    sql = 'DROP TABLE IF EXISTS Factor;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Factor (Period TEXT, Hour TEXT, Value FLOAT);'
    cur.execute(sql)
    factors = list()
    for row in range(1, sh.nrows):
        period = int(sh.cell_value(row, 0))
        hour = int(sh.cell_value(row,1))
        factor = sh.cell_value(row, 3)
        print period, hour, factor
        factors.append((period,hour,factor))
    cur.executemany('INSERT INTO Factor VALUES (?,?,?)', factors)
    conn.commit()

#write shift forward, backward & away to excel
def shift_to_excel():
    writefile = os.getcwd() + '\\' + excel_shift_name
    writer = ExcelWriter(writefile)
    print gdx_file
    zone_dict = dict()
    zone_dict['BEL_Z'] = 'BEL'

    print 'Retrieving shiftaway'
    shiftaway = gdx_to_df(gdx_file, 'shiftaway')
    old_index = shiftaway.index.names
    shiftaway['C'] = [zone_dict[z] for z in shiftaway.index.get_level_values('Z')]
    shiftaway.set_index('C', append=True, inplace=True)
    shiftaway = shiftaway.reorder_levels(['C'] + old_index)
    shiftaway.reset_index(inplace=True)
    shiftaway = pivot_table(shiftaway, 'shiftaway', index=['P','H','Z'], columns=['C'], aggfunc=np.sum)


    print 'Retrieving shiftforward'
    shiftforward = gdx_to_df(gdx_file, 'shiftforwards')
    old_index = shiftforward.index.names
    shiftforward['C'] = [zone_dict[z] for z in shiftforward.index.get_level_values('Z')]
    shiftforward.set_index('C', append=True, inplace=True)
    shiftforward = shiftforward.reorder_levels(['C'] + old_index)
    shiftforward.reset_index(inplace=True)
    shiftforward = pivot_table(shiftforward, 'shiftforwards', index=['P','H','Z'], columns=['C'], aggfunc=np.sum)

    print 'Retrieving shiftbackward'
    shiftbackward = gdx_to_df(gdx_file, 'shiftbackwards')
    old_index = shiftbackward.index.names
    shiftbackward['C'] = [zone_dict[z] for z in shiftbackward.index.get_level_values('Z')]
    shiftbackward.set_index('C', append=True, inplace=True)
    shiftbackward = shiftbackward.reorder_levels(['C'] + old_index)
    shiftbackward.reset_index(inplace=True)
    shiftbackward = pivot_table(shiftbackward, 'shiftbackwards', index=['P','H','Z'], columns=['C'], aggfunc=np.sum)

    # First Merge
    shift = merge(shiftforward, shiftbackward, left_index=True, right_index=True, how='outer', suffixes=['_forward', '_backward'])
    shift = merge(shift, shiftaway, left_index=True, right_index=True, how='outer', suffixes=['', '_away'])

    print 'Writing demand and prices to Excel'
    shift.to_excel(writer, na_rep=0.0, sheet_name=sh_shift_name, merge_cells=False)

    writer.close()

#calculate the compensation factor for each period
def calculate_comp_factor():
    print 'open shifting excel file'
    wbread = openpyxl.load_workbook(excel_shift_name)
    print 'shifting file loaded'
    sheet = wbread.get_sheet_by_name(sh_shift_name)
    print list_compensation
    for i in range(1,amount_of_periods+1):
        forward = 0
        backward = 0
        away = 0
        for j in range(2,length_period+2):
            # print ('j: ', j)
            forward = forward + abs(sheet.cell(row = (i-1)*length_period+j,column = 4).value)
            backward = backward + abs(sheet.cell(row = (i-1)*length_period+j,column = 5).value)
            away = away + abs(sheet.cell(row = (i-1)*length_period+j,column = 6).value)
        print forward
        print backward
        print away
        compensate = away/(forward+backward)
        print 'compensate: ',compensate
        list_compensation[i-1] = list_compensation[i-1]*math.pow(compensate,0.8)
        print 'compensate new value: ', list_compensation[i-1]

# funtion to choose the right demand profiles based on EV penetration (range0-100)
def setRightDemandProfilesEV(percentageEV):
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()
    print os.getcwd()

    row = int(percentageEV/10)+1

    bookref = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\DemR\DemResRef.xlsx"))
    bookmin = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\DemR\DemResMin.xlsx"))
    bookmax = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\DemR\DemResMax.xlsx"))
    bookflat = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\DemR\DemResFlatPrice.xlsx"))
    sqlref = 'DROP TABLE IF EXISTS Dem_ref_profile;'
    sqlmin = 'DROP TABLE IF EXISTS Dem_min_profile;'
    sqlmax = 'DROP TABLE IF EXISTS Dem_max_profile;'
    sqlflat = 'DROP TABLE IF EXISTS Dem_flat_profile;'
    cur.execute(sqlref)
    cur.execute(sqlmin)
    cur.execute(sqlflat)
    cur.execute(sqlmax)
    sqlref = 'CREATE TABLE IF NOT EXISTS Dem_ref_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    sqlmin = 'CREATE TABLE IF NOT EXISTS Dem_min_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    sqlmax = 'CREATE TABLE IF NOT EXISTS Dem_max_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    sqlflat = 'CREATE TABLE IF NOT EXISTS Dem_flat_profile (Season FLOAT, Zone TEXT, Hour TEXT, Demand FLOAT);'
    cur.execute(sqlref)
    cur.execute(sqlmin)
    cur.execute(sqlmax)
    cur.execute(sqlflat)
    demref = list()
    demmin = list()
    demmax = list()
    demflat = list()
    zone = 'BEL_Z'
    amount_of_days = length_period/24
    for season in range (0,4):
        print 'season: ', season
        for day in range(0,amount_of_days):
            if day == startday_weekend or day == startday_weekend+1:
                print 'weekendday with sheet index = ', season*2+1
                shref=bookref.sheet_by_index(season*2+1)
                shmin=bookmin.sheet_by_index(season*2+1)
                shmax=bookmax.sheet_by_index(season*2+1)
                shflat=bookflat.sheet_by_index(season*2+1)
            else:
                print 'weekday with sheet index = ', season*2
                shref=bookref.sheet_by_index(season*2)
                shmin=bookmin.sheet_by_index(season*2)
                shmax=bookmax.sheet_by_index(season*2)
                shflat=bookflat.sheet_by_index(season*2)
            for col in range(1,shref.ncols):
                hour = int(shref.cell_value(0,col)) + 24*day
                valueref = shref.cell_value(row,col)
                valuemin = shmin.cell_value(row,col)
                valuemax = shmax.cell_value(row,col)
                valueflat = shflat.cell_value(row,col)
                demref.append((season+1,zone,hour,valueref))
                demmin.append((season+1,zone,hour,valuemin))
                demmax.append((season+1,zone,hour,valuemax))
                demflat.append((season+1,zone,hour,valueflat))
    cur.executemany('INSERT INTO Dem_ref_profile VALUES (?,?,?,?)', demref)
    cur.executemany('INSERT INTO Dem_min_profile VALUES (?,?,?,?)', demmin)
    cur.executemany('INSERT INTO Dem_max_profile VALUES (?,?,?,?)', demmax)
    cur.executemany('INSERT INTO Dem_flat_profile VALUES (?,?,?,?)', demflat)
    conn.commit()
    print 'Done price profiles'

# function to choose the right elasticity matrix based on EV penetration (range0-100)
def setRightElasticityMatrix(percentageEV):
    print os.getcwd()
    conn = sq.connect("database/database.sqlite")
    cur = conn.cursor()
    print os.getcwd()

    sheet = int(percentageEV/10)

    book = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\Elasticity.xlsx"))
    bookElastRange = xlrd.open_workbook(os.path.join(os.getcwd() , "excel\DemR\ElasticitiesRangeSpring.xlsx"))
    sql = 'DROP TABLE IF EXISTS Elasticity;'
    cur.execute(sql)
    sql = 'CREATE TABLE IF NOT EXISTS Elasticity (Season FLOAT, Hour1 TEXT, Hour2 TEXT, Price_Elasticity FLOAT);'
    cur.execute(sql)
    elasticity = list()
    amount_of_days = length_period/24
    for season in range (0,4):
        if season != season_range:
            print 'season: ', season
            for day in range(0,amount_of_days):
                if day == startday_weekend or day == startday_weekend+1:
                    print 'weekend, normal season, sheet = ', season*2+1
                    sh=book.sheet_by_index(season*2+1)
                else:
                    print 'weekday, normal season, sheet = ', season*2
                    sh=book.sheet_by_index(season*2)
                for row in range(3,sh.nrows):
                    hour1 = int(sh.cell_value(row, 0)) + 24*day
                    for col in range(1,sh.ncols):
                        if col < 13:
                            if row > 14 + col:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day+1)
                            else:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day)
                            if hour2 > length_period:
                                hour2 = hour2 - length_period
                        else:
                            if col > 11 + row-2:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day-1)
                            else:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day)
                            if hour2 < 1:
                                hour2 = hour2 + length_period
                        value = sh.cell_value(row,col)
                        elasticity.append((season+1,hour1,hour2,value))
        else:
            print 'season: ', season
            for day in range(0,amount_of_days):
                if day == startday_weekend or day == startday_weekend+1:
                    print 'weekend, range_season, sheet = ', sheet
                    sh=bookElastRange.sheet_by_index(sheet)
                else:
                    print 'weekday, range_season, sheet = ', sheet
                    sh=bookElastRange.sheet_by_index(sheet)
                for row in range(3,sh.nrows):
                    hour1 = int(sh.cell_value(row, 0)) + 24*day
                    for col in range(1,sh.ncols):
                        if col < 13:
                            if row > 14 + col:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day+1)
                            else:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day)
                            if hour2 > length_period:
                                hour2 = hour2 - length_period
                        else:
                            if col > 11 + row-2:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day-1)
                            else:
                                hour2 = int(sh.cell_value(2, col)) + 24*(day)
                            if hour2 < 1:
                                hour2 = hour2 + length_period
                        value = round(sh.cell_value(row,col),4)
                        elasticity.append((season+1,hour1,hour2,value))
    cur.executemany('INSERT INTO Elasticity VALUES (?,?,?,?)', elasticity)
    conn.commit()
    print 'Done elasticities'

x = 'dit is een test'
y = 5
z = x + str(y)
print z