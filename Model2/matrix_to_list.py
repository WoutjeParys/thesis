__author__ = 'Wout'

import os

from pandas import pivot_table, merge, ExcelWriter, DataFrame
import numpy as np

#import some usefull things
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from gams_addon import gdx_to_df, DomainInfo

from openpyxl.styles import Style, Border, Alignment, Protection, Font, colors

print 'opening this shit'
wbread = openpyxl.load_workbook('excel/auxiliary/matrix.xlsx')
wbwrite = openpyxl.load_workbook('excel/Demand.xlsx')
print 'thing loaded'
print wbread.get_sheet_names()
sheetread_real = wbread.get_sheet_by_name('2015_real')
sheetread_forecast = wbread.get_sheet_by_name('2015_forecast')
sheetwrite = wbwrite.get_sheet_by_name('Sheet1')
countrows = sheetread_real.get_highest_row() #should be 365
countcolumns = sheetread_real.get_highest_column() #should be 97
realempty = 0
forecastempty = 0
for i in range(1,countrows+1):
    print i
    writecell = sheetwrite.cell(row = (i-1)*(countcolumns-1)+1, column = 1)
    writecell.value = sheetread_real.cell(row = i, column = 1).value
    for j in range(2,countcolumns+1):
        # print j
        writecell = sheetwrite.cell(row = (i-1)*(countcolumns-1)+(j-1),column = 2)
        if sheetread_real.cell(row = i, column = j).value < 1:
            # print 'value:'
            # print sheetread_real.cell(row = i, column = j).value
            realempty = realempty + 1
            if sheetread_forecast.cell(row = i, column = j).value < 1:
                forecastempty = forecastempty + 1
                writecell.value = (sheetread_forecast.cell(row = i-7, column = j).value + sheetread_forecast.cell(row = i+7, column = j).value)/2
            else:
                writecell.value = sheetread_forecast.cell(row = i, column = j).value
        else:
            writecell.value = sheetread_real.cell(row = i, column = j).value

print 'empty cells in real:'
print realempty
print 'empty cells in forecast:'
print forecastempty

# for i in range(1,countrows+1):
#     print i
#     # writecell = sheetwrite.cell(row = (i-1)*(countcolumns-1)+1, column = 1)
#     # writecell.value = sheetread.cell(row = i, column = 1).value
#     for j in range(2,countcolumns+1):
#         x = 5
#         # print j
#         writecell = sheetwrite.cell(row = (i-1)*(countcolumns-1)+(j-1),column = 3)
#         writecell.value = sheetread_forecast.cell(row = i, column = j).value

wbwrite.save('excel/Demand.xlsx')
